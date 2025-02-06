import customtkinter as tk
from PIL import Image, ImageTk, ImageDraw
import pickle
import openpyxl as px
from openpyxl.styles import PatternFill, Color


class painter:

    # consts
    # EX_IMG = Image.open("daily duties (web).png")
    # EX_FILE = "painting.xlsx"
    # EX_DATA_FILE = "colored.data"
    # PX = 329
    # PY = 185
    #
    # IM_IMG = Image.open("progress.png")
    # IM_FILE = "progress.png"
    # IM_DATA_FILE = "im_colored.data"
    # IM_PX = 3840
    # IM_PY = 2160

    IMGS = {1: Image.open("daily duties (web).png"), 2: Image.open("imageToColor.jpg")}
    FILES = {1: "painting.xlsx", 2: "progress.png"}
    DATA_FILES = {1: "colored.data", 2: "im_colored.data"}
    PX = {1: 329, 2: 3840}
    PY = {1: 185, 2: 2160}


    # tkinter
    root = tk.CTk()
    root.geometry("500x400")

    default_font = tk.CTkFont(family="Century", size=16)

    name_lbl = tk.CTkLabel(root, font=default_font)
    image_lbl = tk.CTkLabel(root, font=default_font)
    main_frame = tk.CTkFrame(root)
    radio_lbl = tk.CTkLabel(main_frame, text="Выберите режим покраски:", font=default_font)
    mode = tk.IntVar()
    mode.set(1)
    excel_radio_btn = tk.CTkRadioButton(main_frame, text="Excel", variable=mode, value=1, font=default_font)
    image_radio_btn = tk.CTkRadioButton(main_frame, text="Image", variable=mode, value=2, font=default_font)

    input_lbl = tk.CTkLabel(main_frame, text="Введите число для покраски:", font=default_font)
    to_color = tk.StringVar(root)
    input_entry = tk.CTkEntry(main_frame, textvariable=to_color, font=default_font)
    info_frame = tk.CTkFrame(main_frame, fg_color="gray")
    colored_lbl = tk.CTkLabel(info_frame, text="Покрашено:", font=default_font)
    colored_v = tk.IntVar()
    colored_number_lbl = tk.CTkLabel(info_frame, textvariable=colored_v, font=default_font)
    remains_lbl = tk.CTkLabel(info_frame, text="Осталось:", font=default_font)
    remains_v = tk.IntVar()
    remains_number_lbl = tk.CTkLabel(info_frame, textvariable=remains_v, font=default_font)
    count_lbl = tk.CTkLabel(info_frame, text="Всего пикселей:", font=default_font)
    count_number_lbl = tk.CTkLabel(info_frame, font=default_font)

    def __init__(self):
        self.settings()

        self.root.mainloop()

    def settings(self):
        # settings
        self.name_lbl.configure(text=self.IMGS[self.mode.get()].filename)
        imgtk = ImageTk.PhotoImage(self.IMGS[self.mode.get()].resize((400, 226)))
        self.image_lbl.configure(image=imgtk, text="")
        self.image_lbl.image = imgtk

        self.excel_radio_btn.configure(command=lambda: self.change_mode(self.mode.get()))
        self.image_radio_btn.configure(command=lambda: self.change_mode(self.mode.get()))

        self.input_entry.bind("<Return>", lambda event: self.validation(event))
        self.colored_v.set(self.get_colored(self.DATA_FILES[self.mode.get()]))
        all_pixels = self.PX[self.mode.get()] * self.PY[self.mode.get()]
        self.remains_v.set(all_pixels - self.colored_v.get())
        self.count_number_lbl.configure(text=all_pixels)

        # aligns
        self.name_lbl.pack()
        self.image_lbl.pack()
        self.main_frame.pack(padx=20,pady=20)
        self.radio_lbl.grid(row=0, column=0)
        self.excel_radio_btn.grid(row=1, column=0)
        self.image_radio_btn.grid(row=2, column=0)
        self.input_lbl.grid(row=0, column=1)
        self.input_entry.grid(row=1, column=1)
        self.info_frame.grid(row=2, column=1, pady=20)
        self.colored_lbl.grid(row=0, column=0, sticky=tk.W, pady=3)
        self.colored_number_lbl.grid(row=0, column=1, sticky=tk.W)
        self.remains_lbl.grid(row=1, column=0, sticky=tk.W, pady=3)
        self.remains_number_lbl.grid(row=1, column=1, sticky=tk.W)
        self.count_lbl.grid(row=2, column=0, sticky=tk.W, pady=3)
        self.count_number_lbl.grid(row=2, column=1, sticky=tk.W)
        print("Settings are completed")

    def change_mode(self, mode):
        imgtk = ImageTk.PhotoImage(self.IMGS[mode].resize((400, 226)))
        self.image_lbl.configure(image=imgtk)
        self.image_lbl.image = imgtk
        self.colored_v.set(self.get_colored(self.DATA_FILES[self.mode.get()]))
        all_pixels = self.PX[self.mode.get()] * self.PY[self.mode.get()]
        self.remains_v.set(all_pixels - self.colored_v.get())
        self.count_number_lbl.configure(text=all_pixels)
        self.name_lbl.configure(text=self.IMGS[self.mode.get()].filename)

    def get_colored(self, fn):
        f = open(fn, 'rb')
        colored = pickle.load(f)
        f.close()
        return colored

    def set_colored(self, fn, colored):
        f = open(fn, 'wb')
        pickle.dump(colored, f)
        f.close()
        # update variables
        self.remains_v.set(self.remains_v.get() - (colored - self.colored_v.get()))
        self.colored_v.set(colored)

    def validation(self, event=None):
        try:
            count = int(self.to_color.get())
            if self.mode.get() == 1:
                self.colore_pixels_excel(count)
            else:
                self.colore_pixels_image(count)
        except ValueError:
            print("Value is not a number")

    def rgb2hex(self, r, g, b):
        return "{:02x}{:02x}{:02x}".format(r, g, b)

    def colore_pixels_excel(self, count):
        if count < 1:
            print("Value is non positive!")
            return


        # load excel file
        wb = px.load_workbook(filename=self.FILES[self.mode.get()])
        ws = wb['Picture']

        # count = int(input("How many pixels have to be colored?\n"))

        # get colored pixels from file
        colored = self.get_colored(self.DATA_FILES[self.mode.get()])

        print("colored pixels:", colored)

        # check number of colored pixels
        if colored >= self.PX[self.mode.get()] * self.PY[self.mode.get()]:
            count = 0
            print("all pixels have already been colored")

        # start position
        starty = colored // self.PX[self.mode.get()]
        startx = colored % self.PX[self.mode.get()]

        # constants for excel
        basex, basey = 2, 2
        step = 6

        # coloring algorithm
        for i in range(starty, self.PY[self.mode.get()]):
            for j in range(startx, self.PX[self.mode.get()]):
                r, g, b, a = self.IMGS[self.mode.get()].getpixel((basex + step * j, basey + step * i))
                ws.cell(i + 1, j + 1).fill = PatternFill(patternType="solid", fgColor=Color(self.rgb2hex(r, g, b)))
                count -= 1
                colored += 1
                if count == 0: break
            startx = 0
            if count == 0: break

        # save number of colored filex in file
        self.set_colored(self.DATA_FILES[self.mode.get()] ,colored)

        # save Excel file
        wb.save(filename=self.FILES[self.mode.get()])

        print("colored pixels after work:", colored)
        print("Work is finished")

    def colore_pixels_image(self, count):
        if count < 1:
            print("Value is non positive!")
            return

        # create or open image file
        im = Image.open(self.FILES[self.mode.get()])
        draw = ImageDraw.Draw(im)

        # get colored pixels from file
        colored = self.get_colored(self.DATA_FILES[self.mode.get()])

        print("colored pixels:", colored)

        # check number of colored pixels
        if colored >= self.PX[self.mode.get()] * self.PY[self.mode.get()]:
            count = 0
            print("all pixels have already been colored")

        # start position
        starty = colored // self.PX[self.mode.get()]
        startx = colored % self.PX[self.mode.get()]

        # coloring algorithm
        for i in range(starty, self.PY[self.mode.get()]):
            for j in range(startx, self.PX[self.mode.get()]):
                r, g, b= self.IMGS[self.mode.get()].getpixel((j, i))
                draw.polygon([(j,i),(j,i)], fill=(r, g, b, 255), outline=None, width=0)
                count -= 1
                colored += 1
                if count == 0: break
            startx = 0
            if count == 0: break

        # save number of colored filex in file
        self.set_colored(self.DATA_FILES[self.mode.get()], colored)

        # save Excel file
        im.save(self.FILES[self.mode.get()])

        print("colored pixels after work:", colored)
        print("Work is finished")

p = painter()