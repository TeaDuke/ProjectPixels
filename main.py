import openpyxl as px
from PIL import Image
from openpyxl.styles import PatternFill, Color
import pickle
import tkinter as tk

def main():
    root = tk.Tk()
    root.geometry("400x600")

    count_l = tk.Label(root, text="Count of pixels to colore")
    count_l.pack()

    count = tk.StringVar(root)

    entry = tk.Entry(root, textvariable=count)
    entry.pack()
    btn = tk.Button(root, text="colore pixels", command= lambda: validation(count.get()))
    btn.pack(pady=5)
    img = tk.PhotoImage(file="daily duties (web).png")
    panel = tk.Label(root, image=img)
    panel.pack()

    root.mainloop()

def validation(value):
    try:
        count = int(value)
        colore_pixels(count)
    except ValueError:
        print("Value is not a number")

def rgb2hex(r,g,b):
    return "{:02x}{:02x}{:02x}".format(r,g,b)

def colore_pixels(count):
    print(type(count))

    # load excel file
    wb = px.load_workbook(filename="painting.xlsx")
    ws = wb['Picture']
    # load image
    im = Image.open("daily duties (web).png")

    width, height = im.size

    #count = int(input("How many pixels have to be colored?\n"))

    # get colored pixels from file
    f = open("colored.data", 'rb')
    colored = pickle.load(f)
    f.close()

    print("colored pixels:", colored)

    # check number of colored pixels
    if colored == width*height:
        count = 0
        print("all pixels have already been colored")

    # start position
    starty = colored // 329
    startx = colored % 329

    # constants
    basex, basey = 2, 2
    step = 6

    # coloring algorithm
    for i in range(starty, 185):
        for j in range(startx, 329):
            r,g,b,a =im.getpixel((basex + step*j, basey + step*i))
            ws.cell(i+1,j+1).fill = PatternFill(patternType="solid", fgColor=Color(rgb2hex(r,g,b)))
            count-=1; colored+=1
            if count == 0: break
        startx = 0
        if count == 0: break

    # save number of colored filex in file
    f = open("colored.data", 'wb')
    pickle.dump(colored, f)
    f.close()

    # save Excel file
    wb.save(filename="painting.xlsx")

    print("colored pixels after work:", colored)
    print("Work is finished")


if __name__ == "__main__":
    main()